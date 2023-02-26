import openpyxl
import os 

# Open an existing workbook
workbook = openpyxl.load_workbook(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
workbook = openpyxl.Workbook()

# Get the active worksheet
worksheet = workbook.active
worksheet = workbook['Sheet']
# Set the name of column A to "ID"
worksheet.column_dimensions['A'].width = 20
worksheet.column_dimensions['A'].title = "ID"
worksheet.cell(row=1, column=1).value = "ID"

# Set the name of column B to "Name"
worksheet.column_dimensions['B'].width = 20
worksheet.column_dimensions['B'].title = "Weighted_average"
worksheet.cell(row=1, column=2).value = "Weighted_average"
# Fill column A with a sequence of numbers

column_name = worksheet.column_dimensions['B'].title


# Print the name of column A
print(f'Column name: {column_name}')


# Save the workbook
workbook.save(os.path.dirname(__file__) +'/totolotto/Lotto.xlsx')