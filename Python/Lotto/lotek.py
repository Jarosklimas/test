#!/usr/bin/env python
# -*- coding: utf -8 -*-

from lxml import html
from datetime import date
import requests
import xlwt
from bs4 import BeautifulSoup
from requests import get
import xlrd
import pandas as pd
import numpy as np
import random
import matplotlib.pyplot as plt
import pylab
import openpyxl
import os

# ----------------------------------------------------------------------To jest sekcja do pobierania danych oraz zczytywania ----------------------------------------------------------------------

URL = "http://www.mbnet.com.pl/dl.xls"                      # lub mozesz użyć URL2 = "http://www.mbnet.com.pl/dl.txt" #URL3 = "http://www.mbnet.com.pl/bazalosl.zip"
response = requests.get(URL)                                # wtedy należy zastosować r = requests.get(URL2) #b = requests.get(URL3)

lokalizacja = int(input(
                        '''\n\n Podaj miejsce zapisu do pliku dl.xls :  \n 
                                1 ) C:/repozytorium/test/Python/dl.xls 
                                2 ) D:/priv_repo/test/Python/dl.txt \n'''))

if lokalizacja == 1:
    filepath = "C:/repozytorium/test/Python/Lotto/dl.xls"

else:
    filepath = "D:/priv_repo/test/Python/Lotto/dl.xls"       # filepath2 = 'D:/priv_repo/test/Python/dl.txt'
                                                             # filepath3 = "D:/priv_repo/test/Python/bazalosl.zip"
with open(filepath, 'wb') as f:
    f.write(response.content)
    f.close

file_1 = xlwt.Workbook(encoding="utf-8")                     # otwarcie pliku i kodowanie pliku

# --------------------------------------------------------------------------------koniec---------------------------------------------------------------------------------------


def data_from_excel(filepath):
    data_list = []
    data = xlrd.open_workbook(filepath)
    sheet = data.sheet_by_index(0)
    sheet.cell_value(0, 0)

    for i in range(sheet.nrows):
        data_list.append(sheet.cell_value(i, 0))
    print("Tutaj sa dane ", data_list)

if lokalizacja == 1:
    file_1 = xlrd.open_workbook("C:/repozytorium/test/Python/Lotto/dl.xls")
else:
    file_1 = xlrd.open_workbook("D:/priv_repo/test/Python/Lotto/dl.xls")

file_1 = xlrd.open_workbook(filename=filepath)
print("The number of worksheets is", file_1.nsheets)
print("Worksheet name(s):", file_1.sheet_names())
sh = file_1.sheet_by_index(0)
print(sh.name, sh.nrows, sh.ncols)
# print("Cell D30 is", sh.cell_value(rowx=29, colx=3))
# --------------------------------------------------------------------------------------wyświetlenie kolumn w postaci wierszy ------------------------------------------------------------
'''for rowx in range(-6, 0):
    licz_cyfre = 1
    if licz_cyfre <=49:
        cyfra = sh.col_values(rowx)[-12:]
        licz=cyfra.count(licz_cyfre)
        print("Kolumna  to :",rowx, cyfra)
        print("1\n",licz)
        licz_cyfre = licz_cyfre +1'''
# --------------------------------------------------------------------------------------koniec-----------------------------------------------------------------------------------------

ilosc_column = -6
licz_cyfre = 1
y = 0
cunter = []

while ilosc_column <= -1:
    lista = sh.col_values(ilosc_column)[-51:-1]  # ostatnia kolumna w arkuszu
    licz_cyfre = 1
    print("To jest lista \n \n", lista)
    while licz_cyfre <= 49:
        wynik = lista.count(licz_cyfre)
        cunter.append(wynik)
        licz_cyfre = licz_cyfre + 1
        # licz_wynik = list(range(wynik))
        # print("To jest licznik wyniku", licz_cyfre, licz_wynik, "\n")

    ilosc_column = ilosc_column + 1
print("\n\n\n                 **************\tIlosc kolumn równa sie :\t", ilosc_column,
      " Licznik przestaje liczyć ilość wystąpienia poszczególnych liczb **************\n\n\n")

print("Wierszy jest :", len(lista))

# print("Suma wartosci elementów  listy to {0}".format(y))


moja_lista_1 = cunter[:49]
moja_lista_2 = cunter[49:98]
moja_lista_3 = cunter[98:147]
moja_lista_4 = cunter[147:196]
moja_lista_5 = cunter[196:245]
moja_lista_6 = cunter[245:294]
#print("To jest pierwsza kolumna (-6) z ilości wystapienia każdej liczby(1-49) \n",
#      moja_lista_1, "wszystkich liczb jest", len(moja_lista_1))
#print("To jest druga kolumna (-5) z  ilości wystapienia każdej liczby(1-49)\n",
#      moja_lista_2, "wszystkich liczb jest", len(moja_lista_2))
#print("To jest trzecia kolumna (-4) z  ilości wystapienia każdej liczby(1-49)\n",
#      moja_lista_3, "wszystkich liczb jest", len(moja_lista_3))
#print("To jest druga kolumna (-3) z  ilości wystapienia każdej liczby(1-49)\n",
#      moja_lista_4, "wszystkich liczb jest", len(moja_lista_4))
#print("To jest druga kolumna (-2) z  ilości wystapienia każdej liczby(1-49)\n",
#      moja_lista_5, "wszystkich liczb jest", len(moja_lista_5))
#print("To jest druga kolumna (-1) z  ilości wystapienia każdej liczby(1-49)\n",
#      moja_lista_6, "wszystkich liczb jest", len(moja_lista_6))
# ------------------------------------------------Ile razy wystepuje dana liczba w 'X' losowaniach dodwanie wartościz wszystkich poszczególnych kolumn--------------------------------------------------
i = 0
win = []
os_x = []
while i < 49:

    win = moja_lista_1[i]+moja_lista_2[i]+moja_lista_3[i] + \
        moja_lista_4[i]+moja_lista_5[i]+moja_lista_6[i]
    # print("Wszystkich",i+1, "w ostatnich 20 losowaniach jest :", win)
    os_x.append(win)
    i = i+1

print("To jest os X (ilość wystąpienia kazdej liczby od 1 do 49) w ",
      len(lista), "losowaniach :\n", os_x)
# os=list(enumerate(os_x,start=1))
# print(os)
# ------------------------------------------------koniec ile razy wystepuje dana liczba w 'X' losowaniach ----------------------------------------------------------------------------------------------
# ------------------------------------------------Obliczanie maximum -----------------------------------------------------------------------------------------------------------------------------------
print("Najwięcej, bo aż ", max(os_x), "razy  padły",
      os_x.count(max(os_x)), "liczby. Są to :  ")

maksimum = max(os_x)
for index, value in enumerate(os_x, start=1):
    if value == maksimum:
        print('{} - {}'.format(index, value))

# ------------------------------------------------koniec maximum ----------------------------------------------------------------------------------------------------------

# ------------------------------------------------Obliczanie minimum ----------------------------------------------------------------------------------------------------------
print("Najmniej, bo aż ", min(os_x), "razy  padły",
      os_x.count(min(os_x)), "liczby. Są to :  ")
minimum = min(os_x)
for index, value in enumerate(os_x, start=1):
    if value == minimum:
        print('{} - {}'.format(index, value))
# ------------------------------------------------ koniec minimum ----------------------------------------------------------------------------------------------------------
# ------------------------------------------------ Wyswietlenie która liczba ile razy padła  ----------------------------------------------------------------------------------------------------------
jed = max(os_x)-1
while jed > 0:
    print("troszkę wiecej, bo aż ", jed, "razy  padły",
          os_x.count(jed), "liczby. Są to :  ")
    for index, value in enumerate(os_x, start=1):
        if value == jed:
            print('{} - {}'.format(index, value))

    jed = jed - 1

# ------------------------------------------------ koniec wyswietlenia ile razy padła kolejna liczba -------------------------------------------------------------------------------------------------------------
z = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25,
     26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49]


for sheet_name in file_1.sheet_names():
    arkusz = file_1.sheet_by_name(sheet_name)
# ------------------------------------------------ robi macierz z konketynch 20 wierszy ----------------------------------------------------------------------------------------------------------

rx = -(len(lista))
i = 0												# tutaj mozesz podać na kótrym wierszu ma sie zakończyć licząc od dołu w pliku dl
licznik = 3
while i < len(lista):
    wiersz = arkusz.row_values(rx)[2:]
    losowanie = np.array(wiersz)
    print(losowanie)
    if (rx < 1):
        rx = rx + 1
        i = i + 1


# ---------------------------------------------------- koniec---------------------------------------------------------------------------------------------------------------------------------------

wiersz_ost = arkusz.row_values(-1)[2:]

wiersz_ost_data = arkusz.row_values(-1)[-7]
wiersz_przed_osta = arkusz.row_values(-2)[2:]
wiersz_20_od_konca = arkusz.row_values(-20)[2:]

porównanie = list(set(wiersz_ost).intersection(set(wiersz_przed_osta)))
b = [int(i) for i in porównanie]
# print(b)

print("Przedostatnie losowanie to\t\t\t\t\t\t : ", wiersz_przed_osta)
print("Ostatnie losowanie z dnia :", wiersz_ost_data, " \t\t : ", wiersz_ost)

k = 2
while k < 8:
    kule_ostatnie = arkusz.row_values(-1)[k]
    for index, value in enumerate(os_x, start=1):
        if index == kule_ostatnie:
            print('{} - {}'.format(index, value))
            k = k+1

print("Ta sama liczba w dwóch ostatnich losowaniach to  : ", porównanie)
print(max(os_x))


# ----------------------------------------------------------------------To jest sekcja do losowania liczb----------------------------------------------------------------------
def losowanie_liczb():
    liczby = []
    ileliczb = 6
    maksliczba = 49
    i = 0

    while i < ileliczb:
        liczba = random.randint(1, maksliczba)

        # jesli liczba wystepuje 0 razy w liscie "liczby" to dodaje ją do listy liczby
        if liczby.count(liczba) == 0:

            liczby.append(liczba)
            c = list(set(liczby).intersection(set(porównanie)))
            d = list(set(liczby).intersection(set(wiersz_ost)))
            e = list(set(liczby).intersection(set(wiersz_przed_osta)))
            # print(liczby)
            # print(type(b[1]))
            # print(c)
            if c or d or e:  # wyrzuca z wylosowanych liczb ostanie losowane oraz liczby wspólne dla 2 ostatnich losowań
                del liczby[-1]
                
                # print("hello")
                # print(liczby)
                i = i - 1
            i = i + 1
    print("Wylosowane liczby to :", liczby)

    l = 0
    while l <= ileliczb-1:
        kule = liczby[l]
        for index, value in enumerate(os_x, start=1):
            if index == kule:
                print('{} - {}'.format(index, value))
                l = l+1


losowanie_liczb()
plt.bar(z, os_x, color='red')
#plt.bar(k, os_x, color='blue')
plt.show()
vk = openpyxl.Workbook()
sk = vk.active
sk.title = "Lotto"
for index, value in enumerate(os_x, start=1):
    sk.cell(row=index, column=1).value = value
print(sk.title)
vk.save(os.path.dirname(__file__) +"/Lotto.xlsx")
# --------------------------------------------------------------------------------------koniec-----------------------------------------------------------------------------------------

# ------------------------------------------------ robi macierz od 0-48----------------------------------------------------------------------------------------------------------
'''
macierz = np.arange(49).reshape(7, 7)

print(macierz)
'''
# -------------------------------------------------------koniec-------------------------------------------------------------------------------------------------------------------

'''
# tworze konkretną ilość arkuszy
sheet1 = baza.add_sheet("Wyniki")
sheet2 = baza.add_sheet("Dane")
# pobieram dzisiejszą datę
today = date.today()
# format wyświetlanej daty
data = today.strftime('%d-%m-%Y')
# adres strony z horoskopem
urlwyborcza = "http://www.mbnet.com.pl/dl.xls"
urlwyborcza += data

# ścieżka do elementu XPATH
xpathwyborcza = '//*[@a="holder_230"]/div/a[1]/text()'
# pobiera stronę z horoskopem
pagewyborcza = requests.get(urlwyborcza)
# parsuje stronę
tree= html.fromstring(pagewyborcza.text)
# wyszukuje interesujący nas element
textwyborcza = tree.xpath(xpathwyborcza)
# wyświetla wyniki działania

print(data)
print (pagewyborcza)
print (tree)
print (textwyborcza)
'''
