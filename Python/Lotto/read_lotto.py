import csv
import xlrd
import xlwt
import requests
import random
from requests import get
import xlwt
import openpyxl
import os
from datetime import datetime

class Lotto():
    
# Open the CSV file

    def __init__(self):
        self.url = "https://www.multipasko.pl/wyniki-csv.php?f=lotto"
        self.filepath = os.path.dirname(__file__) + '/totolotto/wyniki-lotto.csv'
        self.workbook = openpyxl.Workbook('Lotto.xlsx')
    
    def read_csv (self):
        # Open the CSV file for reading
        with open(self.filepath, 'r') as file:
            # Create a CSV reader object
            reader = csv.reader(file)

            # Read the rows of the CSV file into a list
            rows = list(reader)

        with open(self.filepath, 'w', newline='') as file:
            # Create a CSV writer object
            writer = csv.writer(file)
            
            # Add the new column labels to the first row
            rows[0].extend(['new column 1', 'new column 2', 'new column 3', 'new column 4'])

            # Write all the rows back to the CSV file
            for row in rows:
                writer.writerow(row)
                
                
    def difference_of_number(number):
        
        subtraction_result = number[0]

        for i in range(1, len(number)):
            subtraction_result -= number[i]
        return subtraction_result
           
    def sum_of_number(numbers):
        
        sum_numbers = sum(numbers)
        return sum_numbers
    
    def weighted_average(values):
        average = 0
        weights = [1, 2, 3, 4, 5, 6]
        
        for value, weight in zip(values, weights):
            average += value * weight
        return average / sum(weights)
        
    def geomet_average(lista):   
    
        iloczyn=1
        for liczba in lista:
            # mnozenie kolejnych selected_numbers
            iloczyn *=liczba
        #dlugość listy
        n=len(lista)
        return iloczyn**(1/n)            
    
    def number_lists(self):
        
        with open(self.filepath, 'r') as file:
            # Create a CSV reader object
            reader = csv.reader(file)
            
            next(reader)
            list_with_ball_number_lists = []
            list_with_data_number_lists = []
            # Iterate over the rows in the CSV file
            
            for row in reader:
               
                row_string= str(row)
                row_string = row_string[1:-1]
                balls = row_string.split(';')
                ball_number_list = balls[4:10]
                data = balls[1:4]
                
                for i, ball in enumerate(ball_number_list):
                    ball_number_list[i] = ball.strip("'")
                 
                for i, ball in enumerate(ball_number_list):
                    ball_number_list[i] = int(ball)
                #print(ball_number_list)
                list_with_ball_number_lists.append(ball_number_list)
                
                for i, list_data in enumerate(data):
                    data[i] = int(list_data)
                list_with_data_number_lists.append(data)
                    
        return list_with_ball_number_lists, list_with_data_number_lists
    
    def sorted_number(self):
        
        self.list_with_sorted_numbers_lists = []
        
        for lst in list_with_ball_number_lists:
            sorted_lst=sorted(lst)
            #print(sorted_lst)
            self.list_with_sorted_numbers_lists.append(sorted_lst)
            
        return self.list_with_sorted_numbers_lists
    
    def count_the_difference_of_numbers(self):
        
        self.list_with_difference_numbers = []
        for lst in list_with_ball_number_lists:
            result = Lotto.difference_of_number(lst)
            #print(result)
            self.list_with_difference_numbers.append(result)
            
        return self.list_with_difference_numbers
    
    def count_the_difference_of_numbers_sorted(self):
        
        self.list_with_difference_numbers_sorted = []
        for lst in self.list_with_sorted_numbers_lists:
            result = Lotto.difference_of_number(lst)
            #print(result)
            self.list_with_difference_numbers_sorted.append(result)
            
        return self.list_with_difference_numbers_sorted
        
    def count_the_sum_of_numbres(self):
        
        self.list_with_sum_numbers = []
        for lst in list_with_ball_number_lists:
            result = Lotto.sum_of_number(lst)
            #print(result)
            self.list_with_sum_numbers.append(result)
        return self.list_with_sum_numbers
        
    def count_the_weighted_average(self):
        
        self.list_with_weighted_average = []
        for lst in list_with_ball_number_lists:
            result = Lotto.weighted_average(lst)
            #print(result)
            self.list_with_weighted_average.append(result)
        return self.list_with_weighted_average
    
    def count_the_weighted_average_of_numbers_sorted(self):
        self.list_with_weighted_average_numbers_sorted = []
        for lst in self.list_with_sorted_numbers_lists:
            result = Lotto.weighted_average(lst)
            #print(result)
            self.list_with_weighted_average_numbers_sorted.append(result)
        return self.list_with_weighted_average_numbers_sorted
    
    def count_the_geomet_average(self, list_with_ball_number_lists):
        
        self.list_with_geomet_average = []
        for lst in list_with_ball_number_lists:
            result = Lotto.geomet_average(lst)
            #print(result)
            self.list_with_geomet_average.append(result)
            
        return self.list_with_geomet_average
    
    def write_the_geomet_average(self):
        
        workbook = openpyxl.load_workbook(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        worksheet = workbook.active       
        worksheet.column_dimensions['A'].width = 20
        worksheet.cell(row=1, column=1).value = "Geomet_average"
        for index, value in enumerate(self.list_with_geomet_average, start=2):
            worksheet.cell(row=index, column=1).value = value
        
        workbook.save(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        
    def write_the_weighted_average(self):
        
        workbook = openpyxl.load_workbook(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        worksheet = workbook.active
        worksheet.column_dimensions['B'].width = 20
        worksheet.cell(row=1, column= 2 ).value = "Weighted_average"
        
        for index, value in enumerate(self.list_with_weighted_average, start=2):
            worksheet.cell(row=index, column=2).value = value
        
        workbook.save(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        
    def write_the_weighted_average_numbers_sorted(self):
        
        workbook = openpyxl.load_workbook(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        worksheet = workbook.active
        worksheet.column_dimensions['F'].width = 30
        worksheet.cell(row=1, column= 6 ).value = "Weighted_average number sorted"
        
        for index, value in enumerate(self.list_with_weighted_average_numbers_sorted, start=2):
            worksheet.cell(row=index, column=6).value = value
        
        workbook.save(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
    
    def write_the_sum_of_numbers(self):
        
        workbook = openpyxl.load_workbook(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        worksheet = workbook.active
        worksheet.column_dimensions['C'].width = 20
        worksheet.cell(row=1, column= 3 ).value = "Sum numbers"
        
        for index, value in enumerate(self.list_with_sum_numbers, start=2):
            worksheet.cell(row=index, column=3).value = value
        
        workbook.save(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
    
    def write_the_difference_of_numbers(self):
        
        workbook = openpyxl.load_workbook(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        worksheet = workbook.active
        worksheet.column_dimensions['D'].width = 20
        worksheet.cell(row=1, column= 4 ).value = "Difference numbers"
        
        for index, value in enumerate(self.list_with_difference_numbers, start=2):
            worksheet.cell(row=index, column=4).value = value
        
        workbook.save(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        
    def write_the_difference_of_numbers_sorted(self):
        
        workbook = openpyxl.load_workbook(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
        worksheet = workbook.active
        worksheet.column_dimensions['E'].width = 25
        worksheet.cell(row=1, column= 5 ).value = "Difference numbers sorted"
        
        for index, value in enumerate(self.list_with_difference_numbers_sorted, start=2):
            worksheet.cell(row=index, column=5).value = value
        
        workbook.save(os.path.dirname(__file__) +"/totolotto/Lotto.xlsx")
    
    def download_totolotto(self):
        self.response = requests.get(self.url)
        
        
    def create_csv(self):
        
        with open(self.filepath, 'wb') as file:
            # Create a CSV reader object

            file.write(self.response.content)
            file.close
        
    def current_data(self):
        
        date_list = [datetime.now().day, datetime.now().month, datetime.now().year]
        
        return date_list
    
    def check_path(self):
        if os.path.exists(os.path.dirname(__file__) + '/totolotto/') == False:
            os.mkdir('totolotto')
            
        if os.path.exists(os.path.dirname(__file__) + '/totolotto/wyniki-lotto.csv') == False:
            Lotto.download_totolotto(self)
            Lotto.create_csv(self)
        
    def check_number(self):
        
        """Looking the same numbers in two last draw numbers lotto 

        Returns:
            list: comparison - all the same numbers of the last two draws
        """
        last_numbers_lotto = list_with_ball_number_lists[-1]
        before_last_numbers_lotto = list_with_ball_number_lists[-2]
        self.comparison = list(set(last_numbers_lotto).intersection(set(before_last_numbers_lotto)))
        
        
        
        return self.comparison
    
    def count_all_number_in_sublist(self):
        
        result_data = {}
        search_number = 1
        ile_losowan = int(input("podaj ile wyswietlic list \n"))
        sublist = list_with_ball_number_lists [-ile_losowan :]
        #print(sublist)

        def count_number (search_number):

            count = 0

            for number in sublist:
                count += number.count(search_number)

            return count # Output: 3

        while search_number <= 49:
            
            result_data[search_number] = (count_number(search_number))
            search_number = search_number +1  
            dic = result_data.items()
        print(dic)
        max_value  = max(result_data.values())
        max_balls = [k for k,v in result_data.items() if v == max_value]
        print(f'Maximum value: {max_value}')
        print(f'Ball number for maximum value: {max_balls}')
      

    def losowanie_liczb(self):
        
        selected_numbers = []
        last_numbers_lotto = list_with_ball_number_lists[-1]
        before_last_numbers_lotto = list_with_ball_number_lists[-2]
        ileliczb = 6
        maksliczba = 49
        i = 0

        while i < ileliczb:
            liczba = random.randint(1, maksliczba)

            # jesli liczba wystepuje 0 razy w liscie "selected_numbers" to dodaje ją do listy selected_numbers
            if selected_numbers.count(liczba) == 0:

                selected_numbers.append(liczba)
                c = list(set(selected_numbers).intersection(set(emp1.check_number())))
                d = list(set(selected_numbers).intersection(set(last_numbers_lotto)))
                e = list(set(selected_numbers).intersection(set(before_last_numbers_lotto)))
              
                if c or d or e :  # wyrzuca z wylosowanych liczb ostanie losowane oraz selected_numbers wspólne dla 2 ostatnich losowań
                    
                    del selected_numbers[-1]
                    i = i - 1
                    
                i = i + 1
                
        #print("Wylosowane selected_numbers to :", selected_numbers)
        
        return selected_numbers
    
        
            
emp1 = Lotto()
list_with_data_number_lists = emp1.number_lists()
date_list = emp1.current_data()
last_date_lotto = list_with_data_number_lists[-1]
if date_list == last_date_lotto:
    print("data jest ok ")
else :
    print ('data nie jest ok' )
    print('Pobieranie danych...')
    emp1.download_totolotto()
    print('Zapis do pliku...')
    emp1.create_csv()
emp1.check_path()
list_with_ball_number_lists, list_with_data_number_lists = emp1.number_lists()

last_numbers_lotto = list_with_ball_number_lists[-1]
before_last_numbers_lotto = list_with_ball_number_lists[-2]

emp1.check_number()
print("ostatnie liczby to : ", last_numbers_lotto,"\n przed ostatnie liczby to : ", before_last_numbers_lotto)

    




list_with_sorted_numbers_lists= emp1.sorted_number()
all_weighted_average = emp1.count_the_weighted_average()
all_geomet_average = emp1.count_the_geomet_average(list_with_ball_number_lists)
all_sum_of_numbers = emp1.count_the_sum_of_numbres()
all_difference_of_numbers = emp1.count_the_difference_of_numbers()
all_difference_of_numbers_sorted = emp1.count_the_difference_of_numbers_sorted()
all_weighted_average_number_sorted = emp1.count_the_weighted_average_of_numbers_sorted()


print("Maximum weight_average the last 10 is ", max(all_weighted_average_number_sorted[-10:]))
print("Minimum weight_average the last 10 is ", min(all_weighted_average_number_sorted[-10:]))
print("Maximum sum the last 10 is ", max(all_sum_of_numbers[-10:]))
print("Minimum sum the last 10 is ", min(all_sum_of_numbers[-10:]))
count_all_numbers = emp1.count_all_number_in_sublist()

selected_numbers = emp1.losowanie_liczb()
selected_numbers_sorted = sorted(selected_numbers)
selected_numbers_geomet_average = Lotto.geomet_average(selected_numbers)
selected_numbers_sum = Lotto.sum_of_number(selected_numbers)
selected_numbers_sub = Lotto.difference_of_number(selected_numbers)
selected_numbers_sorted_sub = Lotto.difference_of_number(selected_numbers_sorted)
selected_numbers_weight_average = Lotto.weighted_average(selected_numbers)
selected_numbers_sorted_weight_average = Lotto.weighted_average(selected_numbers_sorted)

selected_numbers = []
while True:
    selected_numbers = emp1.losowanie_liczb()
    if selected_numbers_sum > min(all_sum_of_numbers[-10:]) :
        if selected_numbers_sub > -150 :
            if selected_numbers_geomet_average > 13 :
                if selected_numbers_sorted_sub < -110 :
                    if selected_numbers_sorted_weight_average > 24 and selected_numbers_sorted_weight_average <38 :
                        if selected_numbers_weight_average < 30 :
                            print("Prawidłowe liczby które spełniają wysztkie warunki to :",selected_numbers)
                            break

        
    

print("To jest średnia geometryczna: ",selected_numbers_geomet_average)
print("To jest suma wszystkich liczb: ",selected_numbers_sum)
print("To jest różnica wszystkich liczby nieposortowanych: ",selected_numbers_sub)
print("To jest różnica liczb posortowanych: ",selected_numbers_sorted_sub)
print("To jest średnia ważona nieposortowanych liczby: ", selected_numbers_weight_average)
print("To jest średnia ważona posortowanych liczby: ", selected_numbers_sorted_weight_average)


emp1.write_the_geomet_average()
emp1.write_the_weighted_average()
emp1.write_the_sum_of_numbers()
emp1.write_the_difference_of_numbers()
emp1.write_the_difference_of_numbers_sorted()
emp1.write_the_weighted_average_numbers_sorted()

#if select_numbers
#print("Gotowe. Sprawdź")

#for index, value in enumerate(list_with_ball_number_lists, start=1):
#    print('{} - {}'.format(index, value))
#for lst in list_with_ball_number_lists:
#    sorted_lst=sorted(lst)
#    print(sorted_lst)
#    self.list_with_sorted_numbers.append(sorted_lst)
#    result = Lotto.weighted_average(lst)
#    print(result)
#print(emp1.number_lists())
#emp1.geomet_average(list_with_ball_number_lists[0])
#result = [list(map(Lotto.geomet_average, lst)) for lst in list_with_ball_number_lists]
#print(result)

